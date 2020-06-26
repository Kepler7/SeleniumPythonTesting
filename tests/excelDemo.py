import openpyxl

book = openpyxl.load_workbook("C:\\Users\\deneb\\Desktop\\seleniumPython\\testdata.xlsx")

sheet = book.active
cell = sheet.cell(row=1, column=2)
#print(cell.value)
#sheet.cell(row=2, column=2).value = "Kepler"

#print(sheet['A5'].value)

#print(sheet.max_row)
#print(sheet.max_column)
rows = sheet.max_row
columns = sheet.max_column

#esto imprime todos lo valores en vertical
#for rowFor in range(1, rows + 1):
    #or columnFor in range(1, columns + 1):
        #cellFor = sheet.cell(row=rowFor, column=columnFor)
        #print(cellFor.value)

# esto imprime los datos del row segun el test case que tu quieras
for rowFor in range(1, rows + 1):
    if sheet.cell(row=rowFor, column=1).value == "Testcase2":
        for columnFor in range(2, columns + 1):
            cellFor = sheet.cell(row=rowFor, column=columnFor)
            print(cellFor.value)


Dict = {}

for rowD in range(1, rows + 1):
    if sheet.cell(row=rowD, column=1).value == "Testcase2":
        for columnD in range(2, columns + 1):
            Dict[sheet.cell(row=1, column=columnD).value] = sheet.cell(row=rowD, column=columnD).value

