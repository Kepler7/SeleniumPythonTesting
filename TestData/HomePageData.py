import openpyxl


class HomePageData:
    test_homePage_data = [{"firstName": "Kepler", "lastName": "Velasco", "email": "keph76@gmail.com", "gender": "Male"},
                          {"firstName": "Deneb", "lastName": "Solano", "email": "denebsolano@gmail.com",
                           "gender": "Female"}]

    wa_test_url = {"burgerKing": "https://wa.me/5215513093333", "aeroMexico": "https://wa.me/5215551334000"}

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\deneb\\Desktop\\seleniumPython\\testdata.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
