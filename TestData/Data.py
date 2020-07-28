import openpyxl


class Data:

    #wa_test_url = {"burgerKing": "https://wa.me/5215513093333", "aeroMexico": "https://wa.me/5215551334000"}

    @staticmethod
    def getTestData(test_case_name):
        """
        This method gets the data from a excel file you need to pass the path
        :param test_case_name: the name of the test case that will be in excel file in column 0 and x row number
        :return: data from excel file
        """
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\deneb\\Desktop\\seleniumPython\\testdata.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
